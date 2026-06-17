import streamlit as st
import pandas as pd
import plotly.express as px
import io
from datetime import datetime
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

# ✅ CONFIG
st.set_page_config(layout="wide")

st.title("📊 Bitácora de Producción")

# =========================
# ✅ COLUMNAS BASE
# =========================
columnas = [
    "Item", "Cut_Mch", "Type", "Program", "Ext_Fam", "Int_Fam",
    "Circuit", "KBNLoc", "Wire_Tube_Splice", "Cross_Reference",
    "Description", "Length", "Term_A", "Joint_to_A", "Seal_A", "Strip_A",
    "Term_B", "Joint_to_B", "Seal_B", "Strip_B", "Doubling", "Levels",
    "Demanda", "Inventario"
]

# =========================
# ✅ PLANTILLA
# =========================
plantilla = pd.DataFrame(columns=columnas)

buffer_plantilla = io.BytesIO()

with pd.ExcelWriter(buffer_plantilla, engine="openpyxl") as writer:
    plantilla.to_excel(writer, index=False, sheet_name="Plantilla")

# ✅ BOTÓN DESCARGA
st.sidebar.download_button(
    label="⬇️ Descargar Plantilla",
    data=buffer_plantilla.getvalue(),
    file_name="plantilla_produccion.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# =========================
# 📂 CARGA DE ARCHIVO
# =========================
st.sidebar.subheader("📂 Cargar archivo")

archivo = st.sidebar.file_uploader(
    "Sube archivo de producción",
    type=["xlsx"]
)

# =========================
# ✅ FLUJO PRINCIPAL
# =========================
if archivo is not None:

    df = pd.read_excel(archivo)

    # ✅ LIMPIEZA
    df = df.loc[:, ~df.columns.str.contains("Unnamed")]

    # ✅ BASE
    df_filtrado = df.copy()

    # ✅ EJEMPLO CALCULO
    df_filtrado["Faltante"] = (
        df_filtrado["Demanda"] - df_filtrado["Inventario"]
    ).clip(lower=0)

    # ✅ MOSTRAR TABLA
    st.subheader("📋 Datos cargados")
    st.dataframe(df_filtrado, use_container_width=True)

else:
    st.warning("⚠️ Sube un archivo para comenzar")
    st.stop()

    # ✅ CREAR COLUMNAS
    df["Faltante"] = (
        df["Demanda"] - df["Inventario"]
    ).clip(lower=0)

    df["Qty_Run"] = df["Faltante"]

    df["Prioridad"] = df["Faltante"].apply(
        lambda x: 1 if x > 500 else (2 if x > 0 else 3)
    )

else:
    st.warning("⚠️ Sube un archivo para comenzar")
    st.stop()

# ✅ ESTILO OSCURO INDUSTRIAL
st.markdown(
    """
    <style>
    .main {background-color: #0e1117; color: white;}
    h1, h2, h3 {color: white;}
    </style>
    """,
    unsafe_allow_html=True
)

# ✅ CARGA DE DATOS
df = df.loc[:, ~df.columns.str.contains("Unnamed")]

st.markdown("""
<style>
.titulo-principal {
    font-size: 60px;
    font-weight: 800;
    margin-top: -30px;
    margin-bottom: 10px;
}
</style>

<div class="titulo-principal">
🏭 SISTEMA DE PRODUCCIÓN  CORTE
</div>
""", unsafe_allow_html=True)

import time

if st.sidebar.checkbox("📺 Modo Planta en Vivo"):
    time.sleep(5)
    st.rerun()

# ✅ PANEL LATERAL (TIPO SOFTWARE)
st.sidebar.title("🔧 Filtros")

maquinas = df["Cut_Mch"].dropna().unique()
seleccion = st.sidebar.multiselect("Selecciona máquina:", maquinas)

# ✅ FILTRADO CORRECTO
if seleccion:
    df_filtrado = df[df["Cut_Mch"].isin(seleccion)]
else:
    df_filtrado = df

# ✅ crear base de trabajo
base_df = df_filtrado.copy()

# ✅ cálculo por fila
base_df["Faltante"] = base_df["Demanda"] - base_df["Inventario"]

# ✅ evitar negativos
base_df["Faltante"] = base_df["Faltante"].clip(lower=0)

# ✅ producción por fila
base_df["Qty_Run"] = base_df["Faltante"]

# ✅ prioridad
base_df["Prioridad"] = base_df["Faltante"].apply(
    lambda x: 1 if x > 500 else (2 if x > 0 else 3)
)

# ✅ KPIs TIPO INDUSTRIAL

col1, col2, col3, col4 = st.columns(4)

def tarjeta_kpi(titulo, valor, color):

    st.markdown(
        f"""
        <div style="
            background-color:{color};
            padding:20px;
            border-radius:10px;
            text-align:center;
            color:white;
        ">
            <h4>{titulo}</h4>
            <h1>{valor}</h1>
        </div>
        """,
        unsafe_allow_html=True
    )

# =========================
# ✅ PRODUCCIÓN TOTAL (TU DEFINICIÓN CORRECTA)
# =========================

produccion_total = max(
    df_filtrado["Demanda"].sum() - df_filtrado["Inventario"].sum(),
    0
)

# =========================
# ✅ PRODUCCIÓN TOTAL (FIJA ✅)
# =========================

produccion_total = max(
    df_filtrado["Demanda"].sum() - df_filtrado["Inventario"].sum(),
    0
)

# =========================
# ✅ REGLA 30%
# =========================

df_filtrado["Limite_30"] = df_filtrado["Demanda"] * 0.3

df_filtrado["Puede_Cortar"] = (
    df_filtrado["Inventario"] < df_filtrado["Limite_30"]
)

# =========================
# ✅ PRODUCCIÓN PERMITIDA (ORDEN COMPLETA ✅)
# =========================

produccion_permitida = df_filtrado.loc[
    df_filtrado["Puede_Cortar"], "Qty_Run"
].sum()

produccion_bloqueada = produccion_total - produccion_permitida

# =========================
# ✅ DEBUG (REMUEVE LUEGO)
# =========================

# =========================
# ✅ KPI
# =========================

porcentaje_bloqueado = (
    (produccion_bloqueada / produccion_total) * 100
    if produccion_total > 0 else 0
)

# =========================
# ✅ MOSTRAR KPIs
# =========================

col1, col2, col3, col4 = st.columns(4)

with col1:
    tarjeta_kpi(
        "⚙️ Producción Total",
        f"{produccion_total:,.0f}",
        "#1f77b4"
    )

with col2:
    tarjeta_kpi(
        "✅ Producción Permitida",
        f"{produccion_permitida:,.0f}",
        "#2ca02c"
    )

with col3:
    tarjeta_kpi(
        "⛔ Producción Bloqueada",
        f"{produccion_bloqueada:,.0f}",
        "#d62728"
    )

with col4:
    tarjeta_kpi(
        "📉 Bloqueada (%)",
        f"{porcentaje_bloqueado:.1f}%",
        "#8B0000"
    )
# =========================
# ✅ CALCULAR KPIs (ANTES DE MOSTRAR)
# =========================

demanda_total = df["Demanda"].sum()
inventario_total = df["Inventario"].sum()

variedad = len(df)
	
# =========================
# ✅ FILA DE ABAJO CENTRADA (ALINEADA CON ARRIBA)
# =========================

# =========================# =========================# =========================

col_sp1, col1, col2, col3, col_sp2 = st.columns([2, 3, 3, 3, 2])

with col_sp1:
    st.empty()

with col1:
    tarjeta_kpi("📦 Demanda Total", f"{demanda_total:,.0f}", "#ff7f0e")

with col2:
    tarjeta_kpi("📊 Inventario", f"{inventario_total:,.0f}", "#2ca02c")

with col3:
    tarjeta_kpi("🔥 Variedad", f"{variedad:,.0f}", "#d62728")

with col_sp2:
    st.empty()

# ✅ FILA ABAJO CENTRADA REAL

# =========================

col_sp1, col1, col2, col3, col_sp2 = st.columns([1.5, 2, 2, 2, 1.5])

st.markdown("---")

# ✅ crear columnas
col1, col2 = st.columns(2, gap="large")

# =========================
# IZQUIERDA
# =========================
with col1:
    st.subheader("📊 Distribución por Prioridad (Detalle Ceros)")

    df_chart = df_filtrado.copy()

    df_chart["Categoria"] = df_chart.apply(
        lambda row: (
            "Críticas con demanda" if row["Demanda"] > 0 and row["Inventario"] == 0 else
            "Críticas sin demanda" if row["Demanda"] == 0 and row["Inventario"] == 0 else
            "Media" if row["Prioridad"] == 2 else
            "Estables"
        ),
        axis=1
    )

    conteo = df_chart["Categoria"].value_counts().reset_index()
    conteo.columns = ["Categoria", "Cantidad"]

    orden = [
        "Críticas con demanda",
        "Críticas sin demanda",
        "Media",
        "Estables"
    ]

    conteo["Categoria"] = pd.Categorical(
        conteo["Categoria"],
        categories=orden,
        ordered=True
    )

    conteo = conteo.sort_values("Categoria")

    # ✅ solo dibujar si hay datos
    if not conteo.empty:
        fig1 = px.bar(
            conteo,
            x="Categoria",
            y="Cantidad",
            text="Cantidad",
            color="Categoria",
            color_discrete_map={
                "Críticas con demanda": "#FF0000",
                "Críticas sin demanda": "#FF9999",
                "Media": "#FFD700",
                "Estables": "#008000"
            }
        )

        fig1.update_layout(showlegend=False)
        fig1.update_traces(textposition="outside")

        st.plotly_chart(fig1, use_container_width=True)
    else:
        st.warning("⚠️ No hay datos para mostrar en el gráfico")

st.markdown("---")
# =========================
# DERECHA
# =========================
with col2:
    st.subheader("🔥 Órdenes Críticas con Demanda por Programa")

    # ✅ FILTRO REAL DE CRÍTICAS
    criticos = df_filtrado[
        (df_filtrado["Demanda"] > 0) & (df_filtrado["Inventario"] == 0)
    ].copy()

    # ✅ AGRUPAR POR PROGRAMA
    crit_programa = (
        criticos.groupby("Program")
        .size()
        .sort_values(ascending=False)
        .head(10)
    )

    df_plot = crit_programa.reset_index()
    df_plot.columns = ["Programa", "Órdenes"]

    if not df_plot.empty:
        fig2 = px.bar(
            df_plot,
            x="Programa",
            y="Órdenes",
            text="Órdenes",
            color="Órdenes",
            color_continuous_scale="reds"
        )

        fig2.update_layout(
            xaxis_tickangle=-45,
            showlegend=False
        )

        fig2.update_traces(textposition="outside")

        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.warning("⚠️ No hay críticas reales con demanda")

# =========================
# ✅ FUNCIÓN BAHÍA
# =========================

def asignar_bahia(maquina):
    try:
        num = int(str(maquina).replace("CM", ""))
    except:
        return "Sin asignar"

    if 1 <= num <= 21:
        return "Bahía 1"
    elif 22 <= num <= 45:
        return "Bahía 2"
    elif 46 <= num <= 71:
        return "Bahía 3"
    elif 72 <= num <= 84:
        return "Bahía 4"
    elif 85 <= num <= 99:
        return "Bahía 5"
    else:
        return "Otra"

# =========================
# ✅ CREAR COLUMNA BAHÍA
# =========================

df_filtrado["Bahia"] = df_filtrado["Cut_Mch"].apply(asignar_bahia)

# =========================
# ✅ CRÍTICAS CON DEMANDA
# =========================

criticas = df_filtrado[
    (df_filtrado["Demanda"] > 0) &
    (df_filtrado["Inventario"] == 0)
].copy()

# =========================
# ✅ ÓRDENES CRÍTICAS POR BAHÍA
# =========================

criticas_bahia = (
    criticas
    .groupby("Bahia")
    .size()
    .reset_index(name="Ordenes_Criticas")
)

# =========================
# ✅ DEMANDA Y PRODUCCIÓN
# =========================

carga_bahia = (
    df_filtrado
    .groupby("Bahia")[["Demanda", "Qty_Run"]]
    .sum()
    .reset_index()
)

# =========================
# ✅ DISTRIBUCIÓN EN COLUMNAS
# =========================
col1, col2 = st.columns(2)

# =========================
# ✅ GRÁFICA IZQUIERDA
# =========================
with col1:
    st.subheader("🔥 Órdenes Críticas con Demanda por Bahia")

    fig1 = px.bar(
        criticas_bahia,
        x="Bahia",
        y="Ordenes_Criticas",
        text="Ordenes_Criticas",
        color="Ordenes_Criticas",
        color_continuous_scale="Reds"
    )

    fig1.update_traces(textposition="outside")

    fig1.update_layout(
        showlegend=False,
        xaxis_title="Bahía",
        yaxis_title="Órdenes"
    )

    st.plotly_chart(fig1, use_container_width=True)

# =========================
# ✅ BASE
# =========================
df_base = df.copy()

df_base["Demanda"] = pd.to_numeric(df_base["Demanda"], errors="coerce").fillna(0)
df_base["Inventario"] = pd.to_numeric(df_base["Inventario"], errors="coerce").fillna(0)

# =========================
# ✅ FALTANTE (como Excel)
# =========================
df_base["Faltante"] = df_base["Demanda"] - df_base["Inventario"]

# =========================
# ✅ CONDICIÓN REAL
# =========================
condicion = df_base["Inventario"] <= (df_base["Demanda"] * 0.30)

# =========================
# ✅ PERMITIDA (CORRECTA)
# =========================
df_base["Permitida"] = df_base["Faltante"].where(
    condicion,
    0
)

# =========================
# ✅ KPIs (VALIDACIÓN)
# =========================
total_faltante = df_base["Faltante"].sum()
total_permitida = df_base["Permitida"].sum()

# =========================
# ✅ GRÁFICA DERECHA
# =========================
with col2:
    st.subheader("📊 Demanda vs Producción vs Permitida por Bahía")

    carga_bahia = df_base.groupby("Bahia")[["Demanda", "Faltante", "Permitida"]].sum().reset_index()

    df_plot = carga_bahia.melt(
        id_vars="Bahia",
        value_vars=["Demanda", "Faltante", "Permitida"],
        var_name="Tipo",
        value_name="Cantidad"
    )

    fig2 = px.bar(
        df_plot,
        x="Bahia",
        y="Cantidad",
        color="Tipo",
        barmode="group",
        text="Cantidad",
        color_discrete_map={
            "Demanda": "#1f77b4",
            "Faltante": "#2ca02c",
            "Permitida": "#17becf"
        }
    )

    fig2.update_traces(texttemplate="%{y:,.0f}", textposition="outside")

    fig2.update_layout(
        xaxis_title="Bahía",
        yaxis_title="Cantidad"
    )

    st.plotly_chart(fig2, use_container_width=True)

st.markdown("---")
st.markdown("## 🔥 Top 20 Máquinas con Más Órdenes Críticas")

# =========================
# ✅ CRÍTICAS REALES (CORREGIDO)
# =========================

criticos_reales = df_filtrado[
    (df_filtrado["Inventario"] == 0) &   # ✅ sin inventario
    (df_filtrado["Demanda"] > 0)         # ✅ pero con demanda
].copy()

# =========================
# ✅ AGRUPAR POR MÁQUINA
# =========================

crit_maquinas = (
    criticos_reales
    .groupby("Cut_Mch")
    .size()
    .sort_values(ascending=False)
    .head(20)
)

# =========================
# ✅ FORMATO PARA GRÁFICA
# =========================

df_plot = crit_maquinas.reset_index()
df_plot.columns = ["Máquina", "Órdenes"]

# =========================
# ✅ GRÁFICA
# =========================

fig = px.bar(
    df_plot,
    x="Máquina",
    y="Órdenes",
    text="Órdenes",
    color="Órdenes",
    color_continuous_scale="Blues"
)

fig.update_traces(textposition="outside")

fig.update_layout(
    xaxis_title="Máquina",
    yaxis_title="Órdenes Críticas"
)

st.plotly_chart(fig, use_container_width=True)

# =========================
# ✅ VALIDACIÓN (MUY IMPORTANTE 🔍)
# =========================

st.markdown("---")
st.markdown("## 📋 Plan de Producción Completo")

# ✅ limpiar columnas basura
df_filtrado = df_filtrado.loc[:, ~df_filtrado.columns.str.contains("Unnamed")]
df_filtrado = df_filtrado.drop(columns=["EFICIENCIA"], errors="ignore")

# ✅ asegurar faltante
df_filtrado["Faltante_Real"] = (
    df_filtrado["Demanda"] - df_filtrado["Inventario"]
).clip(lower=0)

def calcular_prioridad(row):

    if row["Demanda"] > 0 and row["Inventario"] == 0:
        return 1

    elif row["Demanda"] > 0 and row["Inventario"] < row["Demanda"]:
        return 2

    else:
        return 3

df_filtrado["Prioridad"] = df_filtrado.apply(calcular_prioridad, axis=1)

df_filtrado["Demanda_Diaria"] = df_filtrado["Demanda"] / 5

df_filtrado["DOH"] = (
    df_filtrado["Inventario"] / df_filtrado["Demanda_Diaria"]
).replace([float("inf"), -float("inf")], 0)

df_filtrado["DOH"] = df_filtrado["DOH"].fillna(0).round(2)

def color_prioridad(val):
    if val == 1:
        return "background-color: red; color: white;"
    elif val == 2:
        return "background-color: orange; color: black;"
    elif val == 3:
        return "background-color: green; color: white;"
    return ""

st.dataframe(
    df_filtrado[[
        "Cut_Mch",
        "Program",
        "Int_Fam",
        "Circuit",
        "Wire_Tube_Splice",
        "Description",
        "Length",
        "Term_A",
        "Term_B",
        "Demanda",
        "Inventario",
        "Qty_Run",
        "Prioridad"
    ]].style.map(
        color_prioridad,
        subset=["Prioridad"]   # ✅ SOLO esta columna se colorea
    ),
    use_container_width=True,
    hide_index=True
)

st.markdown("---")
st.subheader("🔥 Top Órdenes Críticas (Impacto Real)")

top_n = st.slider("Cantidad de órdenes a mostrar:", 5, 50, 10)

criticos = df_filtrado[df_filtrado["Prioridad"] == 1].copy()

criticos["Impacto"] = (
    criticos["Demanda"]
    - criticos["Inventario"]
    + criticos["Qty_Run"]
)

criticos = criticos.sort_values(by="Impacto", ascending=False).head(top_n)

st.dataframe(
    criticos[[
        "Cut_Mch","Program","Int_Fam","Circuit","Qty_Run"
    ]],
    use_container_width=True,
    hide_index=True
)

# =========================
# ✅ CÁLCULOS BASE
# =========================

# =========================
# ✅ CÁLCULOS BASE
# =========================

df_filtrado["Faltante"] = (
    df_filtrado["Demanda"] - df_filtrado["Inventario"]
).clip(lower=0)

df_filtrado["Demanda_Diaria"] = df_filtrado["Demanda"] / 5  # ✅ según tu regla

df_filtrado["DOH"] = (
    df_filtrado["Inventario"] / df_filtrado["Demanda_Diaria"]
).replace([float("inf"), -float("inf")], 0)

df_filtrado["DOH"] = df_filtrado["DOH"].fillna(0)
df_filtrado["DOH"] = df_filtrado["DOH"].round(2)

import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side

# =========================
st.markdown("---")
st.markdown("## 🧾 Generación de Bitácoras")

# =========================
# ✅ SELECTOR
# =========================
min_doh, max_doh = st.slider(
    "Selecciona rango de DOH:",
    0.0, 10.0,
    (0.0, 1.0),
    step=0.1
)

# =========================
# ✅ CABECERA
# =========================
st.subheader(f"📋 Bitácora DOH: {min_doh:.1f} – {max_doh:.1f}")

# =========================
# ✅ FILTRO
# =========================
bitacora = df_filtrado[
    (df_filtrado["DOH"] >= min_doh) &
    (df_filtrado["DOH"] <= max_doh)
].copy()

# ✅ FILTRO NUEVO (QUITAR SIN DEMANDA)
bitacora = bitacora[bitacora["Demanda"] > 0]

# =========================
# ✅ SETUP
# =========================
bitacora["SETUP"] = bitacora.apply(
    lambda row: "-".join(sorted([str(row["Term_A"]), str(row["Term_B"])])),
    axis=1
)

# =========================
# ✅ ORDEN
# =========================
bitacora = bitacora.sort_values(
    by=["Cut_Mch", "SETUP", "Qty_Run"],
    ascending=[True, True, False]
)

# =========================
# ✅ MOSTRAR
# =========================
if bitacora.empty:
    st.warning("⚠️ No hay datos para este rango")

else:
    columnas_dashboard = [
    "Cut_Mch",
    "Program",
    "Int_Fam",
    "Circuit",
    "Wire_Tube_Splice",
    "Description",
    "Length",
    "Term_A",
    "Term_B",
    "Demanda",
    "Inventario",
    "Qty_Run",
    "DOH"
]

st.dataframe(
    bitacora[columnas_dashboard],
    use_container_width=True,
    hide_index=True
)

from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

buffer = io.BytesIO()

with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

    maquinas = bitacora["Cut_Mch"].dropna().unique()

    for maquina in maquinas:

        df_maquina = bitacora[bitacora["Cut_Mch"] == maquina].copy()

        columnas_finales = [
            "Program", "Int_Fam", "Circuit", "KBNLoc",
            "Wire_Tube_Splice", "Cross_Reference",
            "Description", "Length", "Term_A",
            "Seal_A", "Term_B", "Seal_B",
            "Demanda", "Inventario", "Qty_Run", "DOH"
        ]

        df_maquina = df_maquina[columnas_finales]

        # ✅ EXPORTAR
        df_maquina.to_excel(writer, sheet_name=str(maquina), index=False, startrow=4)

        ws = writer.book[str(maquina)]

        # ✅ TÍTULO (DESPUÉS del to_excel)
        ws["A1"] = "BITÁCORA DE PRODUCCIÓN"
        ws["A1"].alignment = Alignment(horizontal="left")
        ws["A1"].font = Font(size=14, bold=True)

        # ✅ INFO
        ws["A2"] = "Fecha: " + str(pd.Timestamp.today().date())
        ws["A3"] = "Máquina: " + str(maquina)

        # ✅ FREEZE
        ws.freeze_panes = "A6"

        # =========================
        # ✅ FORMATO TABLA
        # =========================
        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        fill_header = PatternFill("solid", fgColor="D9D9D9")

        rows = df_maquina.shape[0]
        cols = df_maquina.shape[1]

        # ✅ ENCABEZADOS
        for col in range(1, cols + 1):
            cell = ws.cell(row=5, column=col)
            cell.border = border
            cell.fill = fill_header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ✅ DATOS
        for row in ws.iter_rows(
            min_row=6,
            max_row=rows + 5,
            min_col=1,
            max_col=cols
        ):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ✅ AUTO AJUSTE
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

# ✅ DESCARGA
buffer.seek(0)

st.download_button(
    label="📥 Descargar Bitácora por Máquina",
    data=buffer.getvalue(),
    file_name=f"Bitacora_DOH_{min_doh:.1f}_{max_doh:.1f}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
